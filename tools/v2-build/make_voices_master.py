#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""こどもの声の一次台帳 (あみさん提供の xlsx) を voices_master.json へ変換する。

なぜ要るか:
  2026-08-14 に「声が反映されていない場所がある」と指摘を受けた。実際に突き合わせたら
  台帳 69行 / 24箇所 に対し、地図に出ていたのは 26件 / 11店だった。原因は
  **xlsx を読む経路がそもそも無かった**こと。声は verified_shops.json に手で写した分しか
  無く、写し漏れても誰も気づけなかった。

  ここを台帳駆動にして、build_mapdata.py が毎回この JSON から声を組み立てるようにする。
  対応先の無い行は落とさず SKIP に理由つきで載せる。SKIP にも地図にも無い行が現れたら
  build_mapdata.py が止まる (gate N62)。

本文は1文字も変えない:
  こどもの声は本人の言葉。どれを載せるかはこちらの領分でも、書き換えるのはクライアントの領分。
  「ラーメン美味しい」と「ラーメンがおいしい」は別の子の声なので、似ていても両方残す。
  セルの行頭「・」は箇条書きの記号なので落とす (cell に原文を丸ごと控える)。

usage: python make_voices_master.py
"""
import io
import json
import os
import sys

import openpyxl

# ⚠ stdout の差し替えは main() の中でやる。モジュールの読み込みだけで差し替えると、
# gate.py がこの中の cell_lines() を使うために import した時に gate 側の出力が壊れる
# (2026-08-14 実測: gate が最後の print で I/O operation on closed file で落ちた)。

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "_source", "2026-08-14_中山商店街_こどもの声台帳.xlsx")
OUT = os.path.join(HERE, "voices_master.json")

# 台帳の書き方 → 地図の店名。1つの声を2店に出すものは list で持つ。
# (ダブルエッグは台帳が本店/4丁目を区別していない。今の本番も同じ声を両方に出しているので、
#  その挙動を保つ。分けたくなったらあみさんに聞いてからにする。)
NAME_MAP = {
    "フラワー中山": "フラワー中山",
    "とらの子": "中華レストラン とらの子",
    "BAKERY&BAKE End Roll": "BAKERY&BAKE EndRoll",
    "中山山の神公園": "中山山の神公園",
    "ケーキ　ナオ": "Cake NAO",
    "柏屋": "柏屋",
    "KAYA": "レストラン KAYA",
    "BURB usedclothing": "BURB usedclothing",
    "中山とびのこ公園": "なかやまとびのこ公園",
    "ダブルエッグ": ["Double Egg", "Double Egg4丁目"],
    "たきみち公園": "たきみち公園",
    "中山市民センター": "中山市民センター",
    "ウエルシア": "ウエルシア仙台中山店",
    "坂の上": "中山の坂の上",
    "MIYA BAGLE": "MIYA BAGLE",
}

# 地図に出す先が無い行。**落とすのではなく、理由つきで見えるところに置く。**
# あみさんの回答が来たら地図へ昇格させるか、reason を確定させる。
SKIP = {
    "ヒルズパーク": "地図にスポットが無い。足すかどうかをあみさんに確認中 (2026-08-14)",
    "中山イオン": "地図にスポットが無い。足すかどうかをあみさんに確認中 (2026-08-14)",
    "中山一丁目公園": "地図にスポットが無い。足すかどうかをあみさんに確認中 (2026-08-14)",
    "とびのこハウス": "地図にスポットが無い。足すかどうかをあみさんに確認中 (2026-08-14)",
    "ダルマ薬局": "台帳の声も「いまはないけど」で閉店を示している。あみさんに確認中 (2026-08-14)",
    "児童館の近くの坂のてっぺん": "「中山の坂の上」と同じ地点かをあみさんに確認中 (2026-08-14)",
    "街道市": "店でなくイベント。扱いをあみさんに確認中 (2026-08-14)",
    "商店街全体": "特定の店でない。入口の紹介文に置く案であみさんに確認中 (2026-08-14)",
    "MIYA BAGLE": "台帳に声が0行 (店は地図に未掲載)",
}


def cell_lines(value):
    """セルを行に割る。行頭の箇条書き記号だけ落とし、本文は触らない。"""
    out = []
    for raw in str(value or "").split("\n"):
        line = raw.strip().strip("　")
        if line.startswith("・"):
            line = line[1:].strip().strip("　")
        if line:
            out.append(line)
    return out


def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8",
                                  errors="replace", line_buffering=True)
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]

    places, total = [], 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        name = str(row[0].value or "").strip()
        if not name:
            continue
        cell = row[1].value
        lines = cell_lines(cell)
        total += len(lines)

        entry = {"master_name": name, "row": row[0].row,
                 "voices": [{"text": t} for t in lines],
                 "cell": str(cell or "")}
        if name in SKIP:
            entry["skip_reason"] = SKIP[name]
        elif name in NAME_MAP:
            target = NAME_MAP[name]
            entry["shops"] = target if isinstance(target, list) else [target]
        else:
            raise SystemExit(
                "台帳の「%s」が NAME_MAP にも SKIP にも無い。"
                "どちらかに載せてから回すこと (黙って落とさないための停止)。" % name)
        places.append(entry)

    doc = {
        "_readme": "こどもの声の一次台帳。本文は1文字も変えない。"
                   "地図の声はここから生成する (build_mapdata.py)。",
        "source": {
            "file": "tools/v2-build/_source/2026-08-14_中山商店街_こどもの声台帳.xlsx",
            "provided_by": "あみさん (宮城大学 事業構想学群 宮﨑ゼミ / 中山商店街振興組合)",
            "received": "2026-08-14",
            "collected": "2026年6月12日",
            "channel": "LINE",
        },
        "counts": {"places": len(places), "voice_lines": total,
                   "mapped_places": sum(1 for p in places if p.get("shops")),
                   "skipped_places": sum(1 for p in places if p.get("skip_reason"))},
        "places": places,
    }
    with io.open(OUT, "w", encoding="utf-8", newline="\n") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)
        f.write("\n")

    print("wrote %s" % OUT)
    print("箇所 %d / 声 %d行 (地図へ %d箇所 / skip %d箇所)"
          % (doc["counts"]["places"], doc["counts"]["voice_lines"],
             doc["counts"]["mapped_places"], doc["counts"]["skipped_places"]))
    for p in places:
        if p.get("skip_reason"):
            print("  skip %-16s %d行  %s" % (p["master_name"], len(p["voices"]),
                                             p["skip_reason"]))


if __name__ == "__main__":
    main()
